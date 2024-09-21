from requests import *
from time import *
from openpyxl import *
import webbrowser
from random import *
try:
    book=load_workbook('store.xlsx')
except FileNotFoundError:
    book=Workbook('store.xlsx')
    book.save('store.xlsx')
    sheet=book.active
book=load_workbook('store.xlsx')
sheet=book.active

while(True):
    try:
        if(sheet['AP1'].value==None or sheet['AP2'].value==None or sheet['AP3'].value==None or sheet['AP4'].value==None):
            ts=input("""
    Go to create thingspeak accont, two channel and 8 feilds for each,
    Than copy write api key of both channels 
    And copy read api key and channel ID of both channels.
    Create now Y/N:""")
            if(ts.lower()=='y'):

                # URL to open
                url = "https://thingspeak.com/"

                # Open the URL in the default web browser
                webbrowser.open(url)
                
            if(sheet['AP1'].value==None):
                wr2_api=input("Please enter write API key of YOUR channel:")
                sheet['AP1']=wr2_api
                book.save('store.xlsx')
            if(sheet['AP4'].value==None):
                wr1_api=input("Please enter write API key of FRIEND channel:")
                sheet['AP4']=wr1_api
                book.save('store.xlsx')
            if(sheet['AP2'].value==None):
                rr_api=input("Please enter read API key of FRIEND channel:")
                sheet['AP2']=rr_api
                book.save('store.xlsx')            
            if(sheet['AP3'].value==None):
                rr_id=input("Please enter ID of FRIEND channel:")
                sheet['AP3']=rr_id
                book.save('store.xlsx')
        else:
            break
    except ConnectionError:
        print("Please connect internet")
        sleep(2)
while(sheet['NN1'].value==None):
    yes2=get(f'https://api.thingspeak.com/update?api_key={sheet['AP1'].value}&field1=1000&field2=1000&field3=1000&field4=1000&field5=1000&field6=1000&field7=1000&field8=1000')
    yes1=get(f'https://api.thingspeak.com/update?api_key={sheet['AP4'].value}&field1=1000&field2=1000&field3=1000&field4=1000&field5=1000&field6=1000&field7=1000&field8=1000')
    if(yes1.status_code==200 and yes1.status_code==200):
        sheet['NN1']=200
        book.save('store.xlsx')
if(sheet['SD1'].value==None):
    sheet['SD1']=time()
    book.save('store.xlsx')
if(sheet['NM1'].value==None):
    us_name=input("Please enter your name:")
    sheet['NM1']=us_name
    book.save('store.xlsx')
print(f"{sheet['NM1'].value}\nN.B: Message is only 8 word")
if(sheet['RN1'].value==None):
    sheet['RN1']=3
    sheet['AL1']=3
    book.save('store.xlsx')
def ran():
    v=[sheet['RN1'].value,sheet['AL1'].value,sheet['AL2'].value]
    r=randrange(1,101)
    while(r in v):
        r=randrange(1,101)
    return r
ot=0      
while(True):
    try:
        def sand():        
            t1=['q','w','e','r','t','y','u','i','o','p','a','s','d','f','g','h','j','k','l','z','x','c','v','b','n','m']
            dt1={}
            n=0
            for m in range(26):
                if(m<10):
                    n=30+m
                    dt1.update({n:t1[m]})
                else:
                    n=m
                    dt1.update({n:t1[m]})

            def number(letter,dt1):
                for m in dt1:
                    if(dt1[m]==letter.lower()):
                        return m
            word=[]
            wordlist=''
            user=input(f">>")+' '
            for m in user:
                if(m.isalpha()):
                    wordlist+=str(number(m,dt1))

                elif(m.isspace()):
                    word.append(wordlist)
                    wordlist=''
            while(True):
                if('' in word):
                    word.remove('')
                if('' not in word):
                    break
            #MPC
            #print(word)
            message=[1000,1000,1000,1000,1000,1000,1000,1000]
            n=0
            if(len(word)!=0):
                for m in word:
                    if(n<=7):
                        message[n]=int(m)
                        n+=1
            #MPC
            #print(message)
            wait=int(time())-sheet['SD1'].value
            if(wait<16):
                second=16-wait
                sleep(second)
            m1=get(f'https://api.thingspeak.com/update?api_key={sheet['AP1'].value}&field1={message[0]}&field2={message[1]}&field3={message[2]}&field4={message[3]}&field5={message[4]}&field6={message[5]}&field7={message[6]}&field8={message[7]}')
            if(m1.status_code==200 and len(word)!=0):
                print(strftime("Message complete sand\tTime: %H:%M:%S %p  %A %d-%B-20%y"))
            elif(len(word)==0):
                print()
            else:
                print(strftime("Message fail to sand\tTime: %H:%M:%S %p  %A %d-%B-20%y"))
            sheet['SD1']=int(time())
            book.save('store.xlsx')
        def receive(ids):
            t1=['q','w','e','r','t','y','u','i','o','p','a','s','d','f','g','h','j','k','l','z','x','c','v','b','n','m']
            dt1={}
            n=0
            for m in range(26):
                if(m<10):
                    n=30+m
                    dt1.update({n:t1[m]})
                else:
                    n=m
                    dt1.update({n:t1[m]})
            #recever
            ids=ids-1
            word2=[]
            receve = get(f'https://api.thingspeak.com/channels/{sheet['AP3'].value}/feeds.json?api_key={sheet['AP2'].value}&results')
            data = receve.json()
            value1 = int(data['feeds'][ids]['field1'])
            if(value1!=1000):
                word2.insert(0,str(value1))
                t=data['feeds'][ids]['created_at']
                t=str(t)
                w=f"{t[t.index("T")+1]}{t[t.index("T")+2]}"
                w=int(w)+3
                if(w==24):
                    w='(+1)T 00'
                elif(w==25):
                    w='(+1)T 01'
                elif(w==26):
                    w='(+1)T 02'
                else:
                    w=f'T {w}'
                h=f"T{t[t.index("T")+1]}{t[t.index("T")+2]}"
                t=t.replace(h,w)
            value2 = int(data['feeds'][ids]['field2'])
            if(value2!=1000):
                word2.insert(1,str(value2))
            value3 = int(data['feeds'][ids]['field3'])
            if(value3!=1000):
                word2.insert(2,str(value3))
            value4 = int(data['feeds'][ids]['field4'])
            if(value4!=1000):
                word2.insert(3,str(value4))
            value5 = int(data['feeds'][ids]['field5'])
            if(value5!=1000):
                word2.insert(4,str(value5))
            value6 = int(data['feeds'][ids]['field6'])
            if(value6!=1000):
                word2.insert(5,str(value6))
            value7 = int(data['feeds'][ids]['field7'])
            if(value7!=1000):
                word2.insert(6,str(value7))
            value8 = int(data['feeds'][ids]['field8'])
            if(value8!=1000):
                word2.insert(7,str(value8))
            list=[]
            p=''
            for m in word2:
                no=0
                for n in m:
                    no+=1
                    p+=n
                    if(no==2):
                        list.append(p)
                        p=''
                        no=0
                list.append('and')
            if(len(list)!=0 and sheet['MMM2'].value!=t):
                list.pop()
                n_w=''
                for m in list:
                    if(m=='and'):
                        n_w+=' '
                    else:
                        n_w+=dt1[int(m)]
                sheet['MMM2']=t
                book.save('store.xlsx')
                return(f"\t\t{"_"*50}\n\t\t{n_w.capitalize()}\n\t\tTime: {t}\n\t\t{"_"*50}")
            else:
                return ' '
        def control():
            receve = get(f'https://api.thingspeak.com/channels/{sheet['AP3'].value}/feeds.json?api_key={sheet['AP2'].value}&results')
            data = receve.json()
            if(sheet['MMM1'].value==None):
                last=int(data['channel']['last_entry_id'])
                sheet['MMM1']=last
                book.save('store.xlsx')
            last=int(data['channel']['last_entry_id'])
            begin=int(sheet['MMM1'].value)
            if(last==0):
                print()
            elif(last!=begin):
                for m in range(begin,last+1):
                    print(receive(m))
                begin=last
                sheet['MMM1']=begin
                book.save('store.xlsx')
            else:
                print()
        if(ot==0):
            sheet['RN1']=ran()
            book.save('store.xlsx')
            sheet['MMM4']=sheet['MMM1'].value
            sheet['MMM3']=sheet['MMM2'].value
            book.save('store.xlsx')
            sheet['AL1']=sheet['RN1'].value
            book.save('store.xlsx')
            ot=1
            sleep(4)
        control()
        sand()
    except ConnectionError:
        print("Please connect internet")
        sleep(2)

